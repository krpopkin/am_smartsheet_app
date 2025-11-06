#!/usr/bin/env python3
"""
Script to apply changes from WIP spreadsheet to the downloaded Smartsheet Excel file.
This creates a modified Excel file that can then be uploaded back to Smartsheet.
"""

import openpyxl
from openpyxl.styles import Font
import os
import glob
import sys
from datetime import datetime
from dotenv import load_dotenv


def get_changes_from_wip(file_path):
    """
    Read the changes from WIP file by comparing Sheet1 and original tabs.
    
    Args:
        file_path: Path to the WIP Excel file
        
    Returns:
        List of dictionaries containing row data and which columns changed
    """
    #print(f"\nReading changes from: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    ws_sheet1 = wb['Sheet1']
    ws_original = wb['original']
    
    # Get header row to map column names to indices
    headers = {}
    for col in range(1, ws_sheet1.max_column + 1):
        header_value = ws_sheet1.cell(1, col).value
        if header_value:
            headers[header_value] = col
    
    #print(f"Headers found: {list(headers.keys())}")
    
    key_col = headers['key']
    changes_list = []
    
    # Compare rows (starting from row 2, since row 1 is header)
    for row_num in range(2, ws_sheet1.max_row + 1):
        key_value = ws_sheet1.cell(row_num, key_col).value
        
        if not key_value:
            continue
        
        # Compare each cell in this row
        changed_columns = {}
        for col_name, col_idx in headers.items():
            sheet1_value = ws_sheet1.cell(row_num, col_idx).value
            original_value = ws_original.cell(row_num, col_idx).value
            
            if sheet1_value != original_value:
                # Store both original and new values
                changed_columns[col_name] = {
                    'original': original_value,
                    'new': sheet1_value
                }
        
        if changed_columns:
            changes_list.append({
                'key': key_value,
                'row_num': row_num,
                'changes': changed_columns
            })
            #print(f"Row {row_num}: Key={key_value}, Changed columns: {list(changed_columns.keys())}")
    
    print(f"\nTotal rows with changes to apply: {len(changes_list)}")
    return changes_list


def apply_changes_to_excel(excel_file_path, changes_list, output_file_path):
    """
    Apply the changes to the downloaded Smartsheet Excel file.
    
    Args:
        excel_file_path: Path to the original downloaded Excel file
        changes_list: List of changes from get_changes_from_wip()
        output_file_path: Path where the updated file should be saved
    """
    #print(f"\nApplying changes to: {excel_file_path}")
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    
    # Assume the data is in the first sheet
    ws = wb.active
    
    # Get header row to map column names to indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(1, col).value
        if header_value:
            headers[header_value] = col
    
    #Add a new column that will contain an "X" to designate changed rows
    header_value = 'Row Updated'
    headers[header_value] = ws.max_column + 1
    
    #print(f"Headers in target Excel file: {list(headers.keys())}")
    
    # We don't need a 'key' column in the target file
    # The 'key' from WIP file is the Smartsheet row number
    # Excel row = Smartsheet row + 1 (because of header)
    
    #print(f"Total rows in target file: {ws.max_row}")
    
    # Apply each change
    changes_applied = 0
    for change_item in changes_list:
        key = change_item['key']
        changes = change_item['changes']
        
        # Calculate the Excel row number: Smartsheet row + 1 for header
        target_row = key + 1
        
        # Verify the row exists
        if target_row > ws.max_row:
            print(f"  ⚠ Warning: Key {key} (Excel row {target_row}) exceeds file rows ({ws.max_row}), skipping")
            continue
        
        #print(f"\n  Updating Excel row {target_row} (Smartsheet row {key}):")
        
        # Apply each column change
        for col_name, value_dict in changes.items():
            if col_name not in headers:
                print(f"    ⚠ Warning: Column '{col_name}' not found in target file, skipping")
                continue
            
            col_idx = headers[col_name]
            new_value = value_dict['new']
            old_value = ws.cell(target_row, col_idx).value
            
            # Update the cell
            cell = ws.cell(target_row, col_idx)
            cell.value = new_value
            
            # Apply red font to highlight the change
            cell.font = Font(color="C00000")
            
            #Mark the row updated column with a value of "X"
            col_idx = headers['Row Updated']
            cell = ws.cell(target_row,col_idx)
            cell.value = "X"
            
            changes_applied += 1
            
            #print(f"    ✓ {col_name}: '{old_value}' → '{new_value}'")
    
    # Save the modified Excel file
    wb.save(output_file_path)
    #print(f"\n{'='*60}")
    print(f"Changes applied: {changes_applied}")
    #print(f"Updated file saved to: {output_file_path}")
    #print(f"{'='*60}")
    
    return True


def main():
    """
    Main function to apply changes from WIP to downloaded Excel file.
    """
    try:
        load_dotenv()
        
        # Configuration
        smartsheet_name = os.getenv("SMARTSHEET_PROJECT_NAME")
        folder_name = f"{smartsheet_name}_program_plan"
        base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
        
        folder_path = f"{base_path}/{folder_name}"
        
        # Find the most recent WIP file
        wip_pattern = os.path.join(folder_path, f"{smartsheet_name}_program_wip*.xlsx")
        wip_files = glob.glob(wip_pattern)
        
        if not wip_files:
            print("Error: No WIP file found")
            sys.exit(1)
        
        wip_file = max(wip_files, key=os.path.getmtime)
        print(f"Using WIP file: {os.path.basename(wip_file)}")
        
        # Find the program plan file (the downloaded Smartsheet Excel file)
        # Pattern: {smartsheet_name}_program_plan_YYYY_MM_DD.xlsx
        plan_pattern = os.path.join(folder_path, f"{smartsheet_name}_program_plan_*.xlsx")
        plan_files = glob.glob(plan_pattern)
        
        # Exclude files that end with "_with_updates.xlsx"
        plan_files = [f for f in plan_files if not f.endswith("_with_updates.xlsx")]
        
        if not plan_files:
            print(f"Error: No program plan file found matching pattern: {smartsheet_name}_program_plan_*.xlsx")
            sys.exit(1)
        
        plan_file = max(plan_files, key=os.path.getmtime)
        print(f"Using program plan file: {os.path.basename(plan_file)}")
        
        # Create output filename
        base_name = os.path.basename(plan_file)
        name_without_ext = os.path.splitext(base_name)[0]
        output_file = os.path.join(folder_path, f"{name_without_ext}_with_updates.xlsx")
        
        # Get list of changes from the WIP file
        changes_list = get_changes_from_wip(wip_file)
        
        if not changes_list:
            print("\nNo changes to apply")
            return
        
        # Apply changes to the Excel file
        success = apply_changes_to_excel(plan_file, changes_list, output_file)
        
        if success:
            print("\n✓ Successfully created updated Excel file!")
            #print(f"Next step: Upload '{os.path.basename(output_file)}' to Smartsheet")
        else:
            print("\n✗ Failed to apply changes")
            sys.exit(1)
        
    except Exception as e:
        print(f"\nERROR: An exception occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()