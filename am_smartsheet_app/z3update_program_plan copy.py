#!/usr/bin/env python3
"""
Script to update Smartsheet by copying from Excel and pasting into Smartsheet.
"""

import openpyxl
import asyncio
import os
import glob
from dotenv import load_dotenv
from datetime import datetime, timedelta
import pandas as pd 

def get_changes_from_wip(file_path):   
    
    df = pd.read_excel(file_path)
    changes_list = sorted(list(df.key.values))
    
    return changes_list


async def update_smartsheet_from_excel(page, excel_file_path, changes_list):
    """
    Update Smartsheet by copying from Excel and pasting into Smartsheet.
    Only updates rows that were changed in the WIP report.  
    """
    print(f"\nReading Excel file: {excel_file_path}")
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    print(f"Columns: {ws.max_column}")
    print(f"Total rows in Excel: {ws.max_row - 1}")
    
    # Navigate to top-left of Smartsheet
    print("\nNavigating to top of Smartsheet...")
    await page.keyboard.press('Control+Home')
    await asyncio.sleep(1)
    
    # Move down one row to skip the header
    await page.keyboard.press('ArrowDown')
    await asyncio.sleep(0.5)
    
    print("\nStarting row-by-row update...")
    
    # Loop through each data row in Excel (starting from row 2)
    for row_num in range(2, ws.max_row + 1):
        
        adj_row_num = row_num - 1 #Excel is always one row higher due to the header row
        
        if adj_row_num <= max(changes_list):
        
            if adj_row_num in changes_list: 
                for col_idx in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row_num, col_idx).value
                    
                    # Copy the cell value to clipboard
                    if cell_value is not None:
                        # Format the value
                        if hasattr(cell_value, 'strftime'):
                            # Date formatting
                            formatted_value = cell_value.strftime('%m/%d/%y')
                        else:
                            formatted_value = str(cell_value)
                        
                        # BUG FIX: Properly escape the string for JavaScript
                        # Replace backslashes first, then quotes
                        formatted_value = formatted_value.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
                        
                        # Use Playwright to copy text to clipboard and paste
                        await page.evaluate(f"navigator.clipboard.writeText('{formatted_value}')")
                        await asyncio.sleep(0.05)
                        await page.keyboard.press('Control+V')
                        await asyncio.sleep(0.1)
                    
                    # Move to next column (except on last column)
                    if col_idx < ws.max_column:
                        await page.keyboard.press('Tab')
                        await asyncio.sleep(0.1)
                
                # Move to next row
                await page.keyboard.press('Home')
                await asyncio.sleep(0.5)
                await page.keyboard.press('ArrowDown')
                await asyncio.sleep(0.5)
                
            else:
                # Skip rows with no updates
                await page.keyboard.press('Home')
                await asyncio.sleep(0.5)
                await page.keyboard.press('ArrowDown')
                await asyncio.sleep(0.5)
                rows_skipped += 1
        
        else:
            break
            
    print(f"  Skipped {rows_skipped} non-milestone rows")
    print(f"  Skipped {rows_skipped_done} milestone rows marked as 'done'")
    
    # Save the Smartsheet
    print("\nSaving Smartsheet...")
    await page.keyboard.press('Control+S')
    await asyncio.sleep(2)


async def main(page):
    """
    Main function to update Smartsheet from Excel file.
    
    Args:
        page: Playwright page object (already logged in to Smartsheet)
    """
    try:
        load_dotenv()
        
        # Configuration
        smartsheet_name = os.getenv("SMARTSHEET_PROJECT_NAME")
        folder_name = f"{smartsheet_name}_program_plan"
        base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
        
        folder_path = f"{base_path}/{folder_name}"
        
        # Find the most recent WIP file
        wip_pattern = os.path.join(folder_path, f"{smartsheet_name}_program_wip*.xlsx")
        wip_files = glob.glob(wip_pattern)
        
        if not wip_files:
            print("Error: No WIP file found")
            sys.exit(1)
        
        wip_file = max(wip_files, key=os.path.getmtime)
        #print(f"Using WIP file: {os.path.basename(wip_file)}")
        
        # Get list of rows changed from the WIP file
        changes_list = get_changes_from_wip(wip_file)
        
        # Find the most recent _with_updates.xlsx file
        pattern = os.path.join(folder_path, f"{smartsheet_name}_program_plan_*_with_updates.xlsx")
        matching_files = glob.glob(pattern)
        
        if not matching_files:
            print("Error: No _with_updates.xlsx file found")
            return
        
        excel_file = max(matching_files, key=os.path.getmtime)
        print(f"Using Excel file: {os.path.basename(excel_file)}")
        
        # Update Smartsheet
        await update_smartsheet_from_excel(page, excel_file, changes_list)
        
        print("\n" + "="*60)
        print("Smartsheet update completed!")
        print("="*60)
        
    except Exception as e:
        print(f"\nERROR: An exception occurred: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    print("This script should be called from the main app with an active Playwright page")
    print("It cannot be run standalone")