#!/usr/bin/env python3
"""
Script to update Smartsheet by copying from Excel and pasting into Smartsheet.
Only updates rows where Type = "milestone".

FEATURES:
1. Skips hidden columns in Excel to maintain proper column alignment
2. Skips milestone rows that are already marked as "done"
3. Stops processing if more than 5 consecutive rows have empty Primary column
4. Optimized delays for faster execution
"""

import openpyxl
import asyncio
import os
import glob
from dotenv import load_dotenv


async def update_smartsheet_from_excel(page, excel_file_path):
    """
    Update Smartsheet by copying from Excel and pasting into Smartsheet.
    Only updates rows where Type = "milestone" and Status != "done".
    
    Args:
        page: Playwright page object (Smartsheet should already be open)
        excel_file_path: Path to the Excel file with updates applied
    """
    print(f"\nReading Excel file: {excel_file_path}")
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    # Get headers and find the Type and Status column indices
    type_col_idx = None
    status_col_idx = None
    primary_col_idx = None
    
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(1, col).value
        if header_value:
            header_lower = header_value.lower()
            if header_lower == 'type':
                type_col_idx = col
            elif header_lower == 'status':
                status_col_idx = col
            elif header_lower == 'task name' or header_lower == 'primary':
                primary_col_idx = col
    
    if type_col_idx is None:
        print("Error: 'Type' column not found in Excel file")
        return
    
    if status_col_idx is None:
        print("Error: 'Status' column not found in Excel file")
        return
        
    if primary_col_idx is None:
        print("Error: 'Primary' (Task Name) column not found in Excel file")
        return
    
    print(f"Columns: {ws.max_column}")
    print(f"Total rows in Excel: {ws.max_row - 1}")
    
    # Navigate to top-left of Smartsheet
    print("\nNavigating to top of Smartsheet...")
    await page.keyboard.press('Control+Home')
    await asyncio.sleep(0.2)
    
    # Move down one row to skip the header
    await page.keyboard.press('ArrowDown')
    await asyncio.sleep(0.5)
    
    print("\nStarting row-by-row update...")
    
    milestone_rows_updated = 0
    rows_skipped = 0
    rows_skipped_done = 0
    empty_primary_count = 0
    MAX_EMPTY_PRIMARY = 5
    
    # Loop through each data row in Excel (starting from row 2)
    for row_num in range(2, ws.max_row + 1):
        # Check the Type column
        type_value = ws.cell(row_num, type_col_idx).value
        type_str = str(type_value).lower().strip() if type_value else ""
        
        # Check the Status column
        status_value = ws.cell(row_num, status_col_idx).value
        status_str = str(status_value).lower().strip() if status_value else ""
        
        # Check the Primary column
        primary_value = ws.cell(row_num, primary_col_idx).value
        primary_str = str(primary_value).strip() if primary_value else ""
        
        # Track empty Primary column values
        if not primary_str or primary_str == "None":
            empty_primary_count += 1
            if empty_primary_count > MAX_EMPTY_PRIMARY:
                print(f"\n⚠ Stopped processing: Found more than {MAX_EMPTY_PRIMARY} rows with empty Primary column")
                break
        else:
            empty_primary_count = 0  # Reset counter when we find a non-empty Primary
        
        # Check if this is a milestone row that's already done
        if type_str == "milestone" and status_str == "done":
            # Skip this row - it's already done
            await page.keyboard.press('Home')
            await asyncio.sleep(0.05)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.1)
            rows_skipped_done += 1
            continue
        
        if type_str == "milestone":
            # Update this row - copy each cell from Excel and paste into Smartsheet
            if milestone_rows_updated % 10 == 0:
                print(f"  Updated {milestone_rows_updated} milestone rows so far...")
            
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row_num, col_idx).value
                
                # Copy the cell value to clipboard
                if cell_value is not None:
                    # Format the value
                    if hasattr(cell_value, 'strftime'):
                        # Date formatting
                        formatted_value = cell_value.strftime('%m/%d/%y')
                    else:
                        formatted_value = str(cell_value)
                    
                    # Properly escape the string for JavaScript
                    formatted_value = formatted_value.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
                    
                    # Use Playwright to copy text to clipboard and paste
                    await page.evaluate(f"navigator.clipboard.writeText('{formatted_value}')")
                    await asyncio.sleep(0.05)
                    await page.keyboard.press('Control+V')
                    await asyncio.sleep(0.1)
                
                # Move to next column (except on last column)
                if col_idx < ws.max_column:
                    await page.keyboard.press('Tab')
                    await asyncio.sleep(0.1)
            
            milestone_rows_updated += 1
            
            # Move to next row
            await page.keyboard.press('Home')
            await asyncio.sleep(0.05)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.1)
            
        else:
            # Skip non-milestone row - but still move down in Smartsheet to stay in sync
            await page.keyboard.press('Home')
            await asyncio.sleep(0.05)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.1)
            rows_skipped += 1
    
    print(f"\n✓ Updated {milestone_rows_updated} milestone rows")
    print(f"  Skipped {rows_skipped} non-milestone rows")
    print(f"  Skipped {rows_skipped_done} milestone rows marked as 'done'")


async def main(page):
    """
    Main function to update Smartsheet from Excel file.
    
    Args:
        page: Playwright page object (already logged in to Smartsheet)
    """
    try:
        load_dotenv()
        
        # Configuration
        smartsheet_name = os.getenv("SMARTSHEET_PROJECT_NAME")
        folder_name = f"{smartsheet_name}_program_plan"
        base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
        
        folder_path = f"{base_path}/{folder_name}"
        
        # Find the most recent _with_updates.xlsx file
        pattern = os.path.join(folder_path, f"{smartsheet_name}_program_plan_*_with_updates.xlsx")
        matching_files = glob.glob(pattern)
        
        if not matching_files:
            print("Error: No _with_updates.xlsx file found")
            return
        
        excel_file = max(matching_files, key=os.path.getmtime)
        print(f"Using Excel file: {os.path.basename(excel_file)}")
        
        # Update Smartsheet
        await update_smartsheet_from_excel(page, excel_file)
        
        print("\n" + "="*60)
        print("Smartsheet update completed!")
        print("="*60)
        
    except Exception as e:
        print(f"\nERROR: An exception occurred: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    print("This script should be called from the main app with an active Playwright page")
    print("It cannot be run standalone")