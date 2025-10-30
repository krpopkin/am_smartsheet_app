#!/usr/bin/env python3
"""
Script to create WIP (Work In Progress) report from Account Management Program Plan
Filters milestones by release, type, status, and start date criteria
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

def create_wip_report(input_file, output_file, output_filename):
    """
    Create WIP report from program plan spreadsheet
    
    Args:
        input_file: Path to input Excel file
        output_file: Path to output Excel file
    """
    
    # Read the input Excel file
    #print(f"Reading input file: {input_file}")
    df = pd.read_excel(input_file)
    
    # Define filtering criteria
    target_releases = [
        'Market Scaling', 
        'Release 5', 
        'Release 4', 
        'Roadmap 2026 (first half)', 
        'Ground Truth', 
        'Release 6'
    ]
    
    target_types = ['milestone']
    
    # Calculate cutoff date (today + 1 day)
    cutoff_date = datetime.now() + timedelta(days=1)
    
    # Add original row number as index (1-based to match Excel row numbering, +2 for header)
    df['original_row'] = range(2, len(df) + 2)
    
    # Apply filters
    #print("Applying filters...")
    filtered_df = df[
        (df['Release'].isin(target_releases)) &
        (df['Type'].isin(target_types)) &
        (df['Status'] != 'done') &
        (df['Start'] <= cutoff_date)
    ].copy()
    
    #print(f"Filtered {len(filtered_df)} rows from {len(df)} total rows")
    
    # Select and rename columns for output
    output_df = pd.DataFrame({
        'Release': filtered_df['Release'],
        'Swimlane': filtered_df['Swimlane'],
        'Primary': filtered_df['Task Name'],
        'Owner': filtered_df['Owner'],
        'Start': filtered_df['Start'],
        'Finish': filtered_df['Finish'],
        'Status Update': filtered_df['Status Update']
    })
    
    # Format date columns to mm-dd-yyyy before writing to Excel
    date_columns = ['Start', 'Finish']
    for col in date_columns:
        output_df[col] = pd.to_datetime(output_df[col]).dt.strftime('%m-%d-%Y')
    
    # Sort by Owner and Start
    output_df = output_df.sort_values(by=['Owner', 'Start'], ascending=[True, True])
    output_df = output_df.reset_index(drop=True)
    
    # Write to Excel
    #print(f"Writing output file: {output_file}")
    output_df.to_excel(output_file, index=False, engine='openpyxl')
    
    # Now apply formatting using openpyxl
    #print("Applying formatting...")
    wb = load_workbook(output_file)
    ws = wb.active
    
    # Set font for entire sheet (Calibri 11) with Top/Left alignment and Wrap Text
    calibri_11 = Font(name='Calibri', size=11)
    for row in ws.iter_rows():
        for cell in row:
            cell.font = calibri_11
            cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
    
    # Format header row (row 1)
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Adjust column widths for better readability
    column_widths = {
        'A': 13,  # Release
        'B': 18,  # Swimlane
        'C': 50,  # Primary
        'D': 11,  # Owner
        'E': 12,  # Start
        'F': 12,  # Finish
        'G': 46,  # Status Update
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Create a copy of the sheet and name it "original"
    #print("Creating hidden 'original' sheet...")
    ws_copy = wb.copy_worksheet(ws)
    ws_copy.title = 'original'
    ws_copy.sheet_state = 'hidden'
    
    # Save final workbook
    wb.save(output_file)
    print(f"Report created successfully: {output_filename}")
    print(f"Total rows in report: {len(output_df)}")

def main():
    """Main function to run the report generation"""
    # Get today's date
    today = datetime.now()
    
    # Define folder and file names
    folder_name = f"am_program_plan"
    filename = f"am_program_plan_{today.year}_{today.month:02d}_{today.day:02d}.xlsx"
    
    base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
    folder_path = f"{base_path}/{folder_name}"
    
    # Full input file path
    input_file = f"{folder_path}/{filename}"
    
    # Output file in the same folder with matching format
    output_filename = f"am_program_wip_{today.year}_{today.month:02d}_{today.day:02d}.xlsx"
    output_file = f"{folder_path}/{output_filename}"
    
    # Check if input file exists
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' not found")
        print(f"Expected path: {input_file}")
        return
    
    # Create the report
    create_wip_report(input_file, output_file, output_filename)

if __name__ == '__main__':
    main()