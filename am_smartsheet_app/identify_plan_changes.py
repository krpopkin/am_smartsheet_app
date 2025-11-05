#!/usr/bin/env python3
"""
Script to compare Sheet1 and original tabs in am_program_wip spreadsheets
and create a changes tab highlighting differences.
"""

import openpyxl
from openpyxl.styles import Font
import os
import glob
import sys
from datetime import datetime
from dotenv import load_dotenv

def find_most_recent_wip_file(base_path, folder_name, smartsheet_name):
    """
    Find the most recent file starting with '{smartsheet_name}_program_wip' in the specified folder.
    
    Args:
        base_path: Base directory path
        folder_name: Folder name within base path
        
    Returns:
        Full path to the most recent file, or None if not found
    """
    # Normalize the path - remove trailing slashes
    folder_path = base_path.rstrip('/') if not folder_name else f"{base_path.rstrip('/')}/{folder_name}"
    
    if not os.path.exists(folder_path):
        print(f"Error: Folder does not exist: {folder_path}")
        return None
    
    # Find all files matching the pattern (non-recursive)
    pattern = os.path.join(folder_path, f"{smartsheet_name}_program_wip*.xlsx")
    matching_files = glob.glob(pattern)
    
    # If no files found, try recursive search in subdirectories
    if not matching_files:
        pattern_recursive = os.path.join(folder_path, "**", f"{smartsheet_name}_program_wip*.xlsx")
        matching_files = glob.glob(pattern_recursive, recursive=True)
    
    if not matching_files:
        print(f"Error: No files matching the program name WIP file found in {folder_path}")
        return None
    
    # Get the most recent file by modification time
    most_recent = max(matching_files, key=os.path.getmtime)
    
    #print(f"Found {len(matching_files)} matching file(s)")
    #print(f"Using most recent: {os.path.basename(most_recent)}")
    #print(f"Last modified: {datetime.fromtimestamp(os.path.getmtime(most_recent))}")
    
    return most_recent


def compare_and_create_changes(file_path):
    """
    Compare Sheet1 and original tabs, create changes tab with highlighted differences.
    
    Args:
        file_path: Path to the Excel file
    """
    from openpyxl.styles import PatternFill, Alignment
    
    #print(f"\nLoading workbook: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    
    # Check if required sheets exist
    if 'Sheet1' not in wb.sheetnames:
        print("Error: 'Sheet1' not found in workbook")
        sys.exit(1)
    
    if 'original' not in wb.sheetnames:
        print("Error: 'original' sheet not found in workbook")
        sys.exit(1)
    
    ws_sheet1 = wb['Sheet1']
    ws_original = wb['original']
    
    # Remove 'changes' sheet if it already exists
    if 'changes' in wb.sheetnames:
        #print("Removing existing 'changes' sheet")
        wb.remove(wb['changes'])
    
    # Create new changes sheet
    ws_changes = wb.create_sheet('changes')
    
    # Get dimensions
    sheet1_rows = ws_sheet1.max_row
    original_rows = ws_original.max_row
    num_cols = ws_sheet1.max_column
    
    #print(f"\nSheet1 has {sheet1_rows} rows")
    #print(f"Original has {original_rows} rows")
    
    # Copy header row to changes sheet with grey background
    grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    header_row = []
    for col in range(1, num_cols + 1):
        header_value = ws_sheet1.cell(1, col).value
        header_row.append(header_value)
        cell = ws_changes.cell(1, col, header_value)
        cell.fill = grey_fill
        cell.alignment = alignment
    
    changes_row_num = 2
    changes_found = 0
    
    # Compare rows (starting from row 2, since row 1 is header)
    max_rows = max(sheet1_rows, original_rows)
    
    for row_num in range(2, max_rows + 1):
        row_has_changes = False
        changed_cols = []
        
        # Check if row exists in both sheets
        row_in_sheet1 = row_num <= sheet1_rows
        row_in_original = row_num <= original_rows
        
        if row_in_sheet1 and not row_in_original:
            # New row in Sheet1 (not in original)
            row_has_changes = True
            
            # Copy entire row and highlight in blue
            for col in range(1, num_cols + 1):
                cell_value = ws_sheet1.cell(row_num, col).value
                new_cell = ws_changes.cell(changes_row_num, col, cell_value)
                new_cell.font = Font(color="0000FF")  # Blue for new rows
                
        elif not row_in_sheet1 and row_in_original:
            # Row deleted from Sheet1 (exists in original but not Sheet1)
            row_has_changes = True
            
            # Copy entire row from original and highlight in red
            for col in range(1, num_cols + 1):
                cell_value = ws_original.cell(row_num, col).value
                new_cell = ws_changes.cell(changes_row_num, col, cell_value)
                new_cell.font = Font(color="FF0000")  # Red for deleted rows
                
        else:
            # Both rows exist - compare cell by cell
            for col in range(1, num_cols + 1):
                sheet1_value = ws_sheet1.cell(row_num, col).value
                original_value = ws_original.cell(row_num, col).value
                
                if sheet1_value != original_value:
                    row_has_changes = True
                    changed_cols.append(col)
            
            if row_has_changes:
                # Copy row from Sheet1 to changes sheet
                for col in range(1, num_cols + 1):
                    cell_value = ws_sheet1.cell(row_num, col).value
                    new_cell = ws_changes.cell(changes_row_num, col, cell_value)
                    
                    # Highlight changed cells in blue
                    if col in changed_cols:
                        new_cell.font = Font(color="0000FF")  # Blue for changed values
        
        if row_has_changes:
            changes_found += 1
            changes_row_num += 1
    
    print(f"\nTotal changes found: {changes_found}")
    
    # Copy column widths from Sheet1 to changes sheet
    for col_idx in range(1, num_cols + 1):
        col_letter_sheet1 = ws_sheet1.cell(1, col_idx).column_letter
        col_letter_changes = ws_changes.cell(1, col_idx).column_letter
        
        # Get the header name to identify Start and Finish columns
        header_name = ws_sheet1.cell(1, col_idx).value
        
        if header_name in ['Start', 'Finish']:
            # Set Start and Finish columns to width 17
            ws_changes.column_dimensions[col_letter_changes].width = 17
        else:
            # Copy width from Sheet1
            original_width = ws_sheet1.column_dimensions[col_letter_sheet1].width
            ws_changes.column_dimensions[col_letter_changes].width = original_width
    
    # Apply alignment to all cells in changes sheet
    for row in ws_changes.iter_rows(min_row=1, max_row=ws_changes.max_row, min_col=1, max_col=num_cols):
        for cell in row:
            cell.alignment = alignment
    
    # Freeze the top row
    ws_changes.freeze_panes = 'A2'
    
    # Save the workbook to the same file (overwrite)
    wb.save(file_path)
    print(f"\nChanges tab added to workbook")
    
    return file_path


def main():
    """Main function to run the comparison."""
    
    try:
        #print("=" * 80)
        #print("Upload Program Changes - WIP File Comparison Tool")
        #print("=" * 80)
        
        load_dotenv()
                
        # Configuration
        smartsheet_name = os.getenv("SMARTSHEET_PROJECT_NAME")
        folder_name = f"{smartsheet_name}_program_plan"
        base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
        
        # Find the most recent WIP file
        wip_file = find_most_recent_wip_file(base_path, folder_name, smartsheet_name)
        
        if wip_file is None:
            print("\nError: No matching file found.")
            sys.exit(1)
        
        # Compare and create changes
        output_file = compare_and_create_changes(wip_file)
        
        #print("\n" + "=" * 80)
        print("Comparison complete!")
        #print(f"Output file: {output_file}")
        #print("=" * 80)
        
    except Exception as e:
        print(f"\nERROR: An exception occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()