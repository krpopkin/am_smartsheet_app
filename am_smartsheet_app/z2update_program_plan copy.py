#!/usr/bin/env python3
"""
Script to update Smartsheet by copying from Excel and pasting into Smartsheet.
Only updates rows where Type = "milestone".

BUGS FIXED:
1. When skipping non-milestone rows, the code wasn't returning to column 1 first,
   causing column misalignment on subsequent milestone rows
2. Single quotes in cell values weren't properly escaped for JavaScript
"""

import openpyxl
import asyncio
import os
import glob
from dotenv import load_dotenv
from datetime import datetime, timedelta

async def update_smartsheet_from_excel(page, excel_file_path):
    """
    Update Smartsheet by copying from Excel and pasting into Smartsheet.
    Only updates rows where Type = "milestone".
    
    Args:
        page: Playwright page object (Smartsheet should already be open)
        excel_file_path: Path to the Excel file with updates applied
    """
    print(f"\nReading Excel file: {excel_file_path}")
    
    # Load the Excel file
    wb = openpyxl.load_workbook(excel_file_path)
    ws = wb.active
    
    # Get headers and find the Type, Start, Primary, and Status column indices
    type_col_idx = None
    start_col_idx = None
    primary_col_idx = None
    status_col_idx = None
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(1, col).value
        if header_value and header_value.lower() == 'type':
            type_col_idx = col
            #break
        if header_value and header_value.lower() == 'start':
            start_col_idx = col
        if header_value and header_value.lower() == 'primary':
            primary_col_idx = col
        if header_value and header_value.lower() == 'status':
            status_col_idx = col
    
    if type_col_idx is None:
        print("Error: 'Type' column not found in Excel file")
        return
    
    if start_col_idx is None:
        print("Error: 'Start' column not found in Excel file")
        return
    
    if primary_col_idx is None:
        print("Error: 'Primary' column not found in Excel file")
        return
    
    if status_col_idx is None:
        print("Error: 'Status' column not found in Excel file")
        return
    
    print(f"Columns: {ws.max_column}")
    print(f"Total rows in Excel: {ws.max_row - 1}")
    
    # Navigate to top-left of Smartsheet
    print("\nNavigating to top of Smartsheet...")
    await page.keyboard.press('Control+Home')
    await asyncio.sleep(1)
    
    # Move down one row to skip the header
    await page.keyboard.press('ArrowDown')
    await asyncio.sleep(0.5)
    
    print("\nStarting row-by-row update...")
    
    milestone_rows_updated = 0
    rows_skipped = 0
    rows_skipped_done = 0
    consecutive_blank_primary = 0
    MAX_EMPTY_PRIMARY = 5
    
    # Calculate cutoff date (today + 1 day)
    cutoff_date = datetime.now() + timedelta(days=1)
    
    # Loop through each data row in Excel (starting from row 2)
    for row_num in range(2, ws.max_row + 1):
        # Check the Primary column in Excel
        primary_value = ws.cell(row_num, primary_col_idx).value
        primary_str = str(primary_value).strip() if primary_value else ""
        
        # Track empty Primary column values
        if not primary_str or primary_str == "None":
            consecutive_blank_primary += 1
            if consecutive_blank_primary > MAX_EMPTY_PRIMARY:
                print(f"\n⚠ Stopping: Found more than {MAX_EMPTY_PRIMARY} consecutive rows with empty Primary column")
                break
        else:
            consecutive_blank_primary = 0  # Reset counter when we find a non-empty Primary
        
        # Check the Type column
        type_value = ws.cell(row_num, type_col_idx).value
        type_str = str(type_value).lower().strip() if type_value else ""
        
        # Check the Status column
        status_value = ws.cell(row_num, status_col_idx).value
        status_str = str(status_value).lower().strip() if status_value else ""
        
        start_value = ws.cell(row_num, start_col_idx).value
        
        # Skip milestone rows that are already marked as "done"
        if type_str == "milestone" and status_str == "done":
            await page.keyboard.press('Home')
            await asyncio.sleep(0.5)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.5)
            rows_skipped_done += 1
            continue
        
        #if type_str == "milestone":
        if type_str == "milestone" and start_value <= cutoff_date:
            # Update this row - copy each cell from Excel and paste into Smartsheet
            if milestone_rows_updated % 10 == 0:
                print(f"  Updated {milestone_rows_updated} milestone rows so far...")
            
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row_num, col_idx).value
                
                # Copy the cell value to clipboard
                if cell_value is not None:
                    # Format the value
                    if hasattr(cell_value, 'strftime'):
                        # Date formatting
                        formatted_value = cell_value.strftime('%m/%d/%y')
                    else:
                        formatted_value = str(cell_value)
                    
                    # BUG FIX: Properly escape the string for JavaScript
                    # Replace backslashes first, then quotes
                    formatted_value = formatted_value.replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n').replace('\r', '')
                    
                    # Use Playwright to copy text to clipboard and paste
                    await page.evaluate(f"navigator.clipboard.writeText('{formatted_value}')")
                    await asyncio.sleep(0.05)
                    await page.keyboard.press('Control+V')
                    await asyncio.sleep(0.1)
                
                # Move to next column (except on last column)
                if col_idx < ws.max_column:
                    await page.keyboard.press('Tab')
                    await asyncio.sleep(0.1)
            
            milestone_rows_updated += 1
            
            # Move to next row
            await page.keyboard.press('Home')
            await asyncio.sleep(0.5)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.5)
            
        else:
            # Skip non-milestone row
            # BUG FIX: Must go to column 1 first before moving down!
            # Otherwise we'll be misaligned on the next milestone row
            await page.keyboard.press('Home')
            await asyncio.sleep(0.5)
            await page.keyboard.press('ArrowDown')
            await asyncio.sleep(0.5)
            rows_skipped += 1
    
    print(f"\n✓ Updated {milestone_rows_updated} milestone rows")
    print(f"  Skipped {rows_skipped} non-milestone rows")
    print(f"  Skipped {rows_skipped_done} milestone rows marked as 'done'")
    
    # Save the Smartsheet
    print("\nSaving Smartsheet...")
    await page.keyboard.press('Control+S')
    await asyncio.sleep(2)


async def main(page):
    """
    Main function to update Smartsheet from Excel file.
    
    Args:
        page: Playwright page object (already logged in to Smartsheet)
    """
    try:
        load_dotenv()
        
        # Configuration
        smartsheet_name = os.getenv("SMARTSHEET_PROJECT_NAME")
        folder_name = f"{smartsheet_name}_program_plan"
        base_path = "/mnt/c/Users/krpop/Amway Corp/Global Account Management Community - Workspace Core Team - Workspace Core Team/Program Status"
        
        folder_path = f"{base_path}/{folder_name}"
        
        # Find the most recent _with_updates.xlsx file
        pattern = os.path.join(folder_path, f"{smartsheet_name}_program_plan_*_with_updates.xlsx")
        matching_files = glob.glob(pattern)
        
        if not matching_files:
            print("Error: No _with_updates.xlsx file found")
            return
        
        excel_file = max(matching_files, key=os.path.getmtime)
        print(f"Using Excel file: {os.path.basename(excel_file)}")
        
        # Update Smartsheet
        await update_smartsheet_from_excel(page, excel_file)
        
        print("\n" + "="*60)
        print("Smartsheet update completed!")
        print("="*60)
        
    except Exception as e:
        print(f"\nERROR: An exception occurred: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    print("This script should be called from the main app with an active Playwright page")
    print("It cannot be run standalone")